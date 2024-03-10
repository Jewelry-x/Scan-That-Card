from flask import Flask, render_template, request, session, redirect, url_for
from image_text_extractor import *
from card_search_tool import *
from config import SESSION_SECRET
import datetime
import os
from openpyxl import Workbook, load_workbook
import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO
import base64

app = Flask(__name__)
app.secret_key = SESSION_SECRET

options = [
    "Yu-Gi-Oh!",
    "Pokémon",
]


def startworkbook():
    global ws
    global wb
    if os.path.exists("output.xlsx"):
        wb = load_workbook("output.xlsx")

        ws = wb.active

    else:
        wb = Workbook()

        ws = wb.active

        data = [
            ["Name", "Date", "Average Price"],
            ["Water Dragon", "01/03", "28.73"],
            ["Water Dragon", "02/03", "33.43"],
            ["Water Dragon", "03/03", "42.16"],
            ["Water Dragon", "05/03", "40.21"],
            ["Water Dragon", "06/03", "20.27"],
            ["Water Dragon", "08/03", "49.04"],
        ]

        for row_data in data:
            ws.append(row_data)

        wb.save("output.xlsx")


def generate_plot(card_name):
    df = pd.read_excel("output.xlsx")

    condition = df["Name"] == card_name

    filtered_df = df[condition]

    if not filtered_df.empty:
        filtered_df = filtered_df.sort_values(by="Date")

        x = filtered_df["Date"]
        y = filtered_df["Average Price"]

        plt.figure()
        plt.plot(x, y, marker="o", linestyle="-")
        plt.xlabel("Date")
        plt.ylabel("Average Price")
        plt.title("Cost over time")
        plt.grid(True)

        img = BytesIO()
        plt.savefig(img, format="png")
        img.seek(0)
        plot_url = base64.b64encode(img.getvalue()).decode()
    else:
        print("No data matching the condition.")

    return plot_url


@app.route("/", methods=["POST"])
def upload_file():
    cards = None

    if request.method == "POST":
        if "file" not in request.files:
            return "No file part"

        file = request.files["file"]

        if file.filename == "":
            return "No selected file"

        if file:
            filename = file.filename
            file.save(filename)

            card_name = extract(filename)
            game_id = get_game_id(request.form["game_select"])
            cards = find_card_price(game_id, card_name)

            session["cards"] = cards

            return redirect(url_for("card"))

    # Render the upload form template
    return render_template("home.html")


@app.route("/")
def home():
    selected_game = options[0]
    return render_template("home.html", options=options, selected_game=selected_game)


@app.route("/card")
def card():
    current_date = datetime.datetime.now().date()
    current_date = current_date.strftime("%d/%m")
    print("Current date (day/month):", current_date)

    cards = session["cards"]

    avg_price = 0.0
    idx = 0
    for card in cards:
        avg_price += float(card[1][1:])
        idx += 1

    avg_price /= idx

    startworkbook()

    existing_df = pd.read_excel("output.xlsx")

    # Define the new data you want to append
    new_data = [cards[0][0], str(current_date), str(avg_price)]

    # Check if the new data already exists in the existing data
    if not existing_df.isin([new_data]).all().any():
        # Append the new data only if it doesn't already exist
        ws.append([cards[0][0], str(current_date), str(avg_price)])

        wb.save("output.xlsx")

    return render_template(
        "card_statistics.html",
        cards=session["cards"],
        plot_url=generate_plot(cards[0][0]),
    )


if __name__ == "__main__":
    app.run(debug=True)
